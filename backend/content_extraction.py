"""
Content Extraction Layer
Handles extraction of text and tabular data from CSV, XLSX, and PDF files.
"""

import csv
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

logger = logging.getLogger(__name__)


class ContentExtractor:
    """Extracts normalized text blocks from various file types."""

    @staticmethod
    def extract_csv(file_path: str) -> List[Dict[str, Any]]:
        """
        Extract CSV content as normalized text blocks.
        
        Args:
            file_path: Path to CSV file
            
        Returns:
            List of text blocks with metadata
        """
        blocks = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return blocks
                
                for row_idx, row in enumerate(reader, start=2):  # Start at 2 (header is row 1)
                    # Create normalized text from row
                    text_parts = []
                    for col_name, col_value in row.items():
                        if col_value:
                            text_parts.append(f"{col_name}: {col_value}")
                    
                    text = " | ".join(text_parts)
                    
                    if text.strip():
                        blocks.append({
                            "text": text,
                            "metadata": {
                                "file_type": "csv",
                                "row_number": row_idx,
                                "columns": list(reader.fieldnames)
                            }
                        })
            
            logger.info(f"Extracted {len(blocks)} rows from {file_path}")
        except Exception as e:
            logger.error(f"Error extracting CSV {file_path}: {e}")
        
        return blocks

    @staticmethod
    def extract_xlsx(file_path: str) -> List[Dict[str, Any]]:
        """
        Extract XLSX content as normalized text blocks.
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            List of text blocks with metadata
        """
        if not openpyxl:
            logger.warning("openpyxl not installed, skipping XLSX extraction")
            return []
        
        blocks = []
        try:
            wb = openpyxl.load_workbook(file_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = None
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    # First row is header
                    if row_idx == 1:
                        headers = [str(h).strip() if h else f"Column_{i}" 
                                  for i, h in enumerate(row)]
                        continue
                    
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Create normalized text from row
                    text_parts = []
                    for col_name, col_value in zip(headers, row):
                        if col_value:
                            text_parts.append(f"{col_name}: {col_value}")
                    
                    text = " | ".join(text_parts)
                    
                    if text.strip():
                        blocks.append({
                            "text": text,
                            "metadata": {
                                "file_type": "xlsx",
                                "sheet": sheet_name,
                                "row_number": row_idx,
                                "columns": headers
                            }
                        })
            
            logger.info(f"Extracted {len(blocks)} rows from {file_path}")
        except Exception as e:
            logger.error(f"Error extracting XLSX {file_path}: {e}")
        
        return blocks

    @staticmethod
    def extract_pdf(file_path: str) -> List[Dict[str, Any]]:
        """
        Extract PDF content as normalized text blocks.
        
        Args:
            file_path: Path to PDF file
            
        Returns:
            List of text blocks with metadata
        """
        if not PyPDF2:
            logger.warning("PyPDF2 not installed, skipping PDF extraction")
            return []
        
        blocks = []
        try:
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                
                for page_num, page in enumerate(pdf_reader.pages, start=1):
                    text = page.extract_text()
                    
                    if text.strip():
                        # Split into paragraphs (split by double newline)
                        paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
                        
                        for para_idx, paragraph in enumerate(paragraphs, start=1):
                            blocks.append({
                                "text": paragraph,
                                "metadata": {
                                    "file_type": "pdf",
                                    "page_number": page_num,
                                    "paragraph_index": para_idx
                                }
                            })
            
            logger.info(f"Extracted {len(blocks)} paragraphs from {file_path}")
        except Exception as e:
            logger.error(f"Error extracting PDF {file_path}: {e}")
        
        return blocks

    @staticmethod
    def extract(file_path: str) -> List[Dict[str, Any]]:
        """
        Auto-detect file type and extract content.
        
        Args:
            file_path: Path to file
            
        Returns:
            List of normalized text blocks
        """
        file_path = str(file_path)
        path_obj = Path(file_path)
        suffix = path_obj.suffix.lower()
        
        if suffix == '.csv':
            return ContentExtractor.extract_csv(file_path)
        elif suffix in ['.xlsx', '.xls']:
            return ContentExtractor.extract_xlsx(file_path)
        elif suffix == '.pdf':
            return ContentExtractor.extract_pdf(file_path)
        else:
            logger.warning(f"Unsupported file type: {suffix}")
            return []

    @staticmethod
    def extract_batch(file_paths: List[str]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract content from multiple files.
        
        Args:
            file_paths: List of file paths
            
        Returns:
            Dict mapping file path to content blocks
        """
        results = {}
        for file_path in file_paths:
            results[file_path] = ContentExtractor.extract(file_path)
        return results
