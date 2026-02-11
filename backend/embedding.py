"""
Uses sentence-transformers (all-MiniLM-L6-v2 model)
Converts text into 384-dimensional vectors
Stores vectors in ChromaDB (vector database)
Automatically chunks long documents
Supports .xlsx and .pdf file formats
"""

import os
import re
import logging
from pathlib import Path
from typing import Optional, Tuple
import chromadb
from sentence_transformers import SentenceTransformer
import openpyxl
from pypdf import PdfReader

logger = logging.getLogger(__name__)


def split_camel_case(name: str) -> str:
    """Split CamelCase into separate words."""
    return re.sub(r'(?<!^)(?=[A-Z])', ' ', name)

def normalize_project_name(folder: str) -> str:
    """Convert folder name to human-readable project name."""
    name = folder.replace('_', ' ').replace('-', ' ')
    name = split_camel_case(name)
    return name.strip().title()

def get_project_id(folder: str) -> str:
    """Generate stable machine-readable project ID."""
    return folder.lower().replace('_', '').replace('-', '').strip()


class EmbeddingManager:
    """
    Manages document embedding and vector storage using ChromaDB and Sentence Transformers.
    Handles document processing, chunking, and embedding generation.
    """
    
    def __init__(self):
        """Initialize embedding manager with ChromaDB and Sentence Transformer model."""
        self.chroma_path = Path(os.getenv('CHROMA_PATH', './data/chroma_db'))
        self.data_path = Path(os.getenv('DATA_PATH', './data/materials'))
        self.chunk_size = 500  # Characters per chunk
        self.chunk_overlap = 100  # Overlap between chunks
        
        # Initialize ChromaDB client
        self.chroma_path.mkdir(parents=True, exist_ok=True)
        self.client = chromadb.PersistentClient(path=str(self.chroma_path))
        
        # Get or create collection
        self.collection = self.client.get_or_create_collection(
            name="material_pricing",
            metadata={"hnsw:space": "cosine"}
        )
        
        # Load embedding model (downloads on first run, ~400MB)
        logger.info("Loading Sentence Transformer model...")
        self.model = SentenceTransformer(
            os.getenv('EMBEDDING_MODEL', 'all-MiniLM-L6-v2'),
            device='cpu'  # Use CPU for demo (production could use GPU)
        )
        logger.info("✓ Embedding model loaded")
    
    def _extract_text_from_pdf(self, file_path: Path) -> str:
        """
        Extract text from PDF file.
        Falls back to reading as plain text if PDF parsing fails.
        
        Args:
            file_path: Path to PDF file
            
        Returns:
            Extracted text from PDF
        """
        try:
            text = ""
            reader = PdfReader(file_path)
            for page in reader.pages:
                text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.warning(f"PDF parsing failed, attempting plain text extraction: {str(e)}")
            # Fallback: try reading as plain text
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except Exception as text_error:
                logger.error(f"Error extracting text from PDF {file_path}: {str(text_error)}")
                return ""
    
    def _extract_text_from_xlsx(self, file_path: Path) -> str:
        """
        Extract text from Excel file.
        Falls back to reading as plain text (CSV) if Excel parsing fails.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Extracted text from Excel cells
        """
        try:
            text = ""
            workbook = openpyxl.load_workbook(file_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    if row_text.strip():
                        text += row_text + "\n"
                
                text += "\n"
            
            return text
        except Exception as e:
            logger.warning(f"Excel parsing failed, attempting plain text extraction: {str(e)}")
            # Fallback: try reading as plain text (CSV format)
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except Exception as text_error:
                logger.error(f"Error extracting text from Excel {file_path}: {str(text_error)}")
                return ""
    
    def _extract_text_from_csv(self, file_path: Path) -> str:
        """
        Extract text from CSV file.
        
        Args:
            file_path: Path to CSV file
            
        Returns:
            Extracted text from CSV
        """
        try:
            text = ""
            with open(file_path, 'r', encoding='utf-8') as f:
                # Read all lines and preserve structure
                lines = f.readlines()
                for line in lines:
                    text += line
            return text
        except Exception as e:
            logger.error(f"Error extracting text from CSV {file_path}: {str(e)}")
            return ""

    
    def _chunk_text(self, text: str, file_path: str, project_name: str, project_id: str) -> list[dict]:
        """
        Split text into overlapping chunks for better embedding quality.
        Injects project context into each chunk.
        
        Args:
            text: Full document text
            file_path: Source file path (for metadata)
            project_name: Normalized project name
            project_id: Stable project ID
            
        Returns:
            List of chunk dictionaries with text and metadata
        """
        chunks = []
        
        # Split into paragraphs first
        paragraphs = text.split('\n\n')
        
        current_chunk = ""
        
        # Context header to prepend to each chunk
        context_header = f"Project: {project_name}\nFile: {Path(file_path).name}\n\n"
        
        for paragraph in paragraphs:
            # If adding this paragraph would exceed chunk_size, save current chunk
            # Note: We don't count context_header in chunk length limit check to avoid strict cutting,
            # but usually it's small.
            if len(current_chunk) + len(paragraph) > self.chunk_size and current_chunk:
                
                # Combine context + content for embedding
                full_text = context_header + current_chunk.strip()
                
                chunks.append({
                    'text': full_text,
                    'file_path': file_path,
                    'length': len(full_text),
                    'project_id': project_id,
                    'project_name': project_name
                })
                
                # Create overlap by including last part of previous chunk
                current_chunk = current_chunk[-(self.chunk_overlap):] + "\n\n" + paragraph
            else:
                current_chunk += "\n\n" + paragraph if current_chunk else paragraph
        
        # Add final chunk
        if current_chunk.strip():
            full_text = context_header + current_chunk.strip()
            chunks.append({
                'text': full_text,
                'file_path': file_path,
                'length': len(full_text),
                'project_id': project_id,
                'project_name': project_name
            })
        
        return chunks
    
    def embed_file(self, file_path: str) -> int:
        """
        Process a single file: read, chunk, embed, and store in vector DB.
        Supports .xlsx, .pdf, and .csv files.
        
        Args:
            file_path: Full path to the file to embed
            
        Returns:
            Number of chunks successfully embedded
            
        Raises:
            ValueError: If file doesn't exist or cannot be read
            Exception: If embedding or storage fails
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise ValueError(f"File not found: {file_path}")
        
        # Check for supported file types
        supported_formats = {'.xlsx', '.pdf', '.csv'}
        if file_path.suffix.lower() not in supported_formats:
            logger.warning(f"Skipping unsupported file format: {file_path}")
            return 0
        
        try:
            logger.info(f"Embedding file: {file_path.name} ({file_path.suffix})")
            
            # Extract content based on file type
            if file_path.suffix.lower() == '.pdf':
                content = self._extract_text_from_pdf(file_path)
            elif file_path.suffix.lower() == '.xlsx':
                content = self._extract_text_from_xlsx(file_path)
            elif file_path.suffix.lower() == '.csv':
                content = self._extract_text_from_csv(file_path)
            else:
                return 0
            
            if not content.strip():
                logger.warning(f"File is empty: {file_path}")
                return 0
            
            # Get relative path for cleaner storage
            try:
                # Always resolve paths for comparison
                data_path_abs = self.data_path.resolve()
                file_path_abs = file_path.resolve()
                
                # Calculate relative path
                if file_path_abs.is_relative_to(data_path_abs):
                    relative_path = file_path_abs.relative_to(data_path_abs)
                else:
                    # Fallback if not under data_path
                    relative_path = Path(file_path.name)
                
                # Normalize to forward slashes and strip redundant 'materials/' if it somehow got in
                rel_str = str(relative_path).replace('\\', '/')
                if rel_str.startswith('materials/'):
                    rel_str = rel_str[len('materials/'):]
                
                normalized_path = rel_str
                
            except Exception as e:
                logger.warning(f"Path normalization failed for {file_path}: {e}")
                normalized_path = file_path.name.replace('\\', '/')
            
            # Remove old embeddings for this file using normalized path
            try:
                self.collection.delete(
                    where={"file_path": normalized_path}
                )
                logger.info(f"Cleared old embeddings for: {normalized_path}")
            except Exception as e:
                logger.debug(f"No previous embeddings to clear: {e}")
            
            # Chunk the document with project context
            folder_name = file_path.parent.name
            project_name = normalize_project_name(folder_name)
            project_id = get_project_id(folder_name)
            
            chunks = self._chunk_text(content, normalized_path, project_name, project_id)
            
            if not chunks:
                logger.warning(f"No chunks created from: {file_path}")
                return 0
            
            # Generate embeddings
            texts = [chunk['text'] for chunk in chunks]
            logger.info(f"Generating embeddings for {len(chunks)} chunks (Project: {project_name})...")
            
            embeddings = self.model.encode(texts, show_progress_bar=False).tolist()
            
            # Prepare documents for storage
            ids = [f"{normalized_path}_{i}" for i in range(len(chunks))]
            metadatas = [
                {
                    "file_path": chunk['file_path'],
                    "chunk_index": str(i),
                    "chunk_length": str(chunk['length']),
                    "project_id": chunk['project_id'],
                    "project_name": chunk['project_name']
                }
                for i, chunk in enumerate(chunks)
            ]
            
            # Store in ChromaDB
            self.collection.add(
                ids=ids,
                embeddings=embeddings,
                documents=texts,
                metadatas=metadatas
            )
            
            logger.info(f"✓ Successfully embedded {len(chunks)} chunks from {file_path.name}")
            return len(chunks)
            
        except Exception as e:
            logger.error(f"Error embedding file {file_path}: {str(e)}", exc_info=True)
            raise
    
    def embed_directory(self, directory: Optional[str] = None) -> dict:
        """
        Recursively embed all supported files (.xlsx, .pdf, .csv) in a directory.
        
        Args:
            directory: Directory path (defaults to DATA_PATH)
            
        Returns:
            Dictionary with embedding statistics
        """
        if directory is None:
            directory = self.data_path
        else:
            directory = Path(directory)
        
        if not directory.exists():
            logger.warning(f"Directory does not exist: {directory}")
            return {"status": "error", "message": f"Directory not found: {directory}"}
        
        logger.info(f"Starting directory embedding: {directory}")
        
        # Find all supported file types
        pdf_files = list(directory.rglob('*.pdf'))
        xlsx_files = list(directory.rglob('*.xlsx'))
        csv_files = list(directory.rglob('*.csv'))
        files = pdf_files + xlsx_files + csv_files
        
        if not files:
            logger.warning(f"No supported files found in {directory}")
            return {
                "status": "success",
                "message": "No files to process",
                "total_files": 0,
                "total_chunks": 0
            }
        
        total_chunks = 0
        successful_files = 0
        failed_files = 0
        
        for file_path in sorted(files):
            try:
                chunks = self.embed_file(str(file_path))
                total_chunks += chunks
                successful_files += 1
            except Exception as e:
                logger.error(f"Failed to embed {file_path}: {str(e)}")
                failed_files += 1
        
        logger.info(
            f"✓ Directory embedding complete. "
            f"Files: {successful_files} successful, {failed_files} failed. "
            f"Total chunks: {total_chunks}"
        )
        
        return {
            "status": "success",
            "total_files": len(files),
            "successful_files": successful_files,
            "failed_files": failed_files,
            "total_chunks": total_chunks
        }
    
    def get_collection_stats(self) -> dict:
        """
        Get statistics about the current collection.
        
        Returns:
            Dictionary with collection statistics
        """
        try:
            count = self.collection.count()
            return {
                "total_documents": count,
                "collection_name": self.collection.name
            }
        except Exception as e:
            logger.error(f"Error getting collection stats: {str(e)}")
            return {"error": str(e)}
    
    def clear_collection(self):
        """
        Clear all embeddings from the collection (useful for reset/testing).
        WARNING: This is a destructive operation.
        """
        try:
            logger.warning(f"Deleting and recreating collection: {self.collection.name}")
            self.client.delete_collection(self.collection.name)
            self.collection = self.client.get_or_create_collection(
                name="material_pricing",
                metadata={"hnsw:space": "cosine"}
            )
            logger.info("✓ Collection cleared and recreated")
        except Exception as e:
            logger.error(f"Error clearing collection: {str(e)}", exc_info=True)
            raise
